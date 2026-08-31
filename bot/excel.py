"""Read the leads table out of an uploaded .xlsx file, exactly as-is."""
from __future__ import annotations

import io
from datetime import date, datetime, time

import openpyxl

from . import config


def _cell_to_str(value) -> str:
    """Match how a human reads the Excel cell, so dedup keys stay stable."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        # Drop a midnight time component the same way Excel display does.
        if value.time() == time(0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


class ExcelParseError(Exception):
    pass


def parse_leads(data: bytes) -> tuple[list[str], list[list[str]]]:
    """Return (header, rows) from the configured source tab.

    Rows are returned in file order with every cell coerced to a string.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ExcelParseError(f"Could not open the file as an Excel workbook: {exc}") from exc

    name = config.SOURCE_SHEET_NAME
    if name and name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        raise ExcelParseError("The sheet is empty.") from None

    header = [_cell_to_str(c) for c in header_raw][: len(config.COLUMNS)]

    rows: list[list[str]] = []
    for raw in rows_iter:
        cells = [_cell_to_str(c) for c in raw][: len(config.COLUMNS)]
        # Pad short rows so every row has the full column count.
        if len(cells) < len(config.COLUMNS):
            cells += [""] * (len(config.COLUMNS) - len(cells))
        if not any(cells):
            continue  # skip fully blank rows
        rows.append(cells)

    wb.close()
    return header, rows
